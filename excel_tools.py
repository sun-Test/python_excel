# -*- coding: utf-8 -*-
'''

@author: sun
'''

import re
import random
from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelUtils(object):
    def __init__(self):
        pass

    @staticmethod
    def firstColumnByReg(sheet, colTitlesRow, regName) -> int:
        for cell in sheet[colTitlesRow]:
            if cell.value:
                #print('firstColumnByReg' + ': ' + cell.value)
                if re.match(pattern=regName, string=str(cell.value)):
                    return cell.column
        return -1

    @staticmethod
    def getColumnValues(sheetRowValues, col) -> list:
        return list(map(lambda row: row[col], sheetRowValues))
        
    @staticmethod
    def filterRowsByColumnReg(sheetRowValues, col, regexFilter) -> list:
        '''
        @param sheetRowValues: openpyxl sheet.values
        @type sheetRowValues: list of tuples
        @param col: column index, that should be filtered
        @type col: integer
        @return: list of OpenpyxlRow tuple, otherwise []
        @rtype: list of OpenpyxlRow tuple
        '''     
        return list(filter(lambda x: re.match(pattern=regexFilter, string=str(x[col])), sheetRowValues))
    
    @staticmethod
    def filterRowsByColumnRegs( sheetRowValues, regexFilters) -> list:
        listOutput = sheetRowValues
        for tupleRegexColumFilter in regexFilters:
            listOutput = ExcelUtils.filterRowsByColumnReg(listOutput, 
                                                          tupleRegexColumFilter[0], 
                                                          tupleRegexColumFilter[1]
                                                          )
        return listOutput
    
    @staticmethod
    def writeColumnToSheet(sheet, startRow, col, dataList):
        for row in range(len(dataList)):
            sheet.cell(row+startRow, col).value = dataList[row]

    @staticmethod
    def writeRowToSheet(sheet, startCol, row, dataList):
        for i in range(len(dataList)):
            sheet.cell(row, startCol + i).value = dataList[i]
   
   
def createExample01():
    wb = Workbook()
    ws = wb.create_sheet('students', 0)
    columnTitles = ['name', 'gender', 'age', 'height']
    ExcelUtils.writeRowToSheet(ws, 3, 1, columnTitles)
    nameList = ['nm' + str(i) for i in range (100, 150)]
    ExcelUtils.writeColumnToSheet(ws, 2, 3, nameList)
    genderList = ['female' if i % 2 == 0 else 'male' for i in range (100, 150)]
    ExcelUtils.writeColumnToSheet(ws, 2, 4, genderList)
    ageList = [random.randrange(5, 10) for i in range(50)]
    ExcelUtils.writeColumnToSheet(ws, 2, 5, ageList)
    heightList = [random.randrange(150, 180) for i in range(50)]
    ExcelUtils.writeColumnToSheet(ws, 2, 6, heightList)

    wb.save('example.xltx')

def test01():
    wb = load_workbook('example.xltx')
    ws = wb['students']
    print(ExcelUtils.firstColumnByReg(ws, 1, r'.{0,3}name'))
    print(ExcelUtils.filterRowsByColumnReg(ws.values, 2, r'.{2,8}0$'))
    print(ExcelUtils.filterRowsByColumnRegs(ws.values, [(2, r'.{2,5}12.$'), (3, r'female')]))

if __name__ == '__main__':
    print('hello')
    #createExample01()
    test01()
